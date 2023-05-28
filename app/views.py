from django.shortcuts import render
import requests
import openai
from openpyxl import Workbook,load_workbook
import csv

openai.api_key = "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"

# Daftar pesan dalam sesi
session_messages = []
# Daftar pencarian sebelumnya
search_history = []

def DISKUSI(request):
    if request.method == "POST":
        prompt = request.POST.get("prompt")
        model_engine = "text-davinci-003"

        completions = openai.Completion.create(
            engine=model_engine,
            prompt=prompt,
            max_tokens=1024,
            n=1,
            temperature=0.5,
        )
        message = completions.choices[0].text

        # Menambahkan prompt dan respon ke dalam daftar pesan sesi
        session_messages.append({"sender": "user", "content": prompt})
        session_messages.append({"sender": "bot", "content": message})

        # Menambahkan pencarian ke dalam daftar history pencarian
        search_history.append(prompt)

        # Simpan ke file Excel
        save_to_excel(prompt, message)

        context = {"messages": session_messages, "searches": search_history}
        return render(request, "index.html", context)
    else:
        context = {"messages": session_messages, "searches": search_history}
        return render(request, "index.html", context)


def newChat(request):
    global session_messages
    session_messages = []
    return render(request, "index.html", {})


def loadChat(request, search):
    global session_messages
    session_messages = []
    prompt = search
    model_engine = "text-davinci-003"

    completions = openai.Completion.create(
        engine=model_engine,
        prompt=prompt,
        max_tokens=1024,
        n=1,
        temperature=0.5,
    )
    message = completions.choices[0].text

    # Menambahkan prompt dan respon ke dalam daftar pesan sesi
    session_messages.append({"sender": "user", "content": prompt})
    session_messages.append({"sender": "bot", "content": message})

    context = {"messages": session_messages, "searches": search_history}
    return render(request, "index.html", context)



# def save_to_csv(prompt, message):
#     with open('chatlog.csv', 'a', newline='') as csvfile:
#         fieldnames = ['Prompt', 'Message']
#         writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

#         # Cek apakah file CSV kosong (belum memiliki header)
#         file_empty = csvfile.tell() == 0

#         if file_empty:
#             writer.writeheader()

#         writer.writerow({'Prompt': prompt, 'Message': message})



def save_to_excel(prompt, message):
    # Buka file Excel atau buat file baru jika belum ada
    try:
        wb = load_workbook('chatlog.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value = 'Prompt'
        sheet.cell(row=1, column=2).value = 'Message'

    # Cari baris terakhir yang sudah terisi
    last_row = sheet.max_row + 1

    # Tambahkan data baru ke baris terakhir
    sheet.cell(row=last_row, column=1).value = prompt
    sheet.cell(row=last_row, column=2).value = message

    # Simpan file Excel
    wb.save('chatlog.xlsx')
